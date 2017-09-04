#! python3
# csv and json data tests


import os, json, csv, openpyxl
from openpyxl.worksheet import Worksheet

#将excel单元格值转换成csv格式
def csvtrans():
    os.chdir('D:\\python\\test\\excelSpreadsheets')
    for excelfile in os.listdir('.'):
        if(not excelfile.endswith('.xlsx')):
            continue
        wb = openpyxl.load_workbook(excelfile)
        for wsheetname in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(wsheetname)
            csvfileobj = open(sheet.title+'.csv', 'w', newline='')
            csvwriter = csv.writer(csvfileobj)
            for rownum in range (1, sheet.max_row+1):
                rowdata = []
                for colnum in range(1, sheet.max_column):
                    rowdata.append(sheet[rownum][colnum].value)
                csvwriter.writerow(rowdata)
            csvfileobj.close()


csvtrans()