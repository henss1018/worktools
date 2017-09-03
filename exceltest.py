import openpyxl, os
from openpyxl.styles import Font

#乘法表
def mutiplicationTable(in_mitrix):
    os.chdir('D:\\python\\test')
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    for i in range(2, int(in_mitrix)+2, 1):
        sheet.cell(row=1, column=i).value = i-1
        sheet.cell(row=i, column=1).value = i-1
    for j in range(2, int(in_mitrix)+2, 1):
        for k in range(2,int(in_mitrix)+2, 1):
            sheet.cell(row=j, column=k).value = int(sheet.cell(row=j, column=1).value
                                                    ) * int(sheet.cell(row=1, column=k).value)
    wb.save('mutiplication_table.xlsx')

#从excel读取二维数组，倒置后存为另一个excel文件；可以考虑直接利用row和culumn参数进行单元格值的倒置，不用中间数组传递
def turnoverTable():
    os.chdir('D:\\python\\test')
    swb = openpyxl.load_workbook('turnovertable.xlsx')
    wb = openpyxl.Workbook()
    ssheet = swb.get_active_sheet()
    sheet = wb.get_sheet_by_name('Sheet')
    i = 0
    j = 0
    #二维数组的初始化方法
    table = [[0 for col in range(ssheet.max_column)] for row in range(ssheet.max_row)]

    for rows in ssheet:
        for cellobj in rows:
            table[i][j] = cellobj.value
            j = j + 1
        i = i + 1
        j = 0

    i = 0
    j = 0
    for i in range(len(table)):
        for j in range(len(table[i])):
            sheet.cell(row=j+1, column=i+1).value = table[i][j]
        i = i + 1
        j = 0
    wb.save('turnovertablenew.xlsx')



#mutiplicationTable(10)
turnoverTable()